import json
import os
import random
import re
import time
import urllib.error
import urllib.parse
import urllib.request

from database.db_handler import get_content_files


class AIQuizGenerationError(Exception):
    pass


class AIQuizGeneratorService:
    ALLOWED_COUNTS = {20, 40, 60, 80, 100}
    MAX_OPTION_WORDS = 5
    MAX_OPTION_CHARS = 80
    CHUNK_SIZE = 24000
    TELEGRAM_DOWNLOAD_LIMIT = 20 * 1024 * 1024

    def __init__(self, api_key=None, api_url=None, model=None):
        # Prefer an explicitly selected provider, then Gemini when its key exists, otherwise Ollama.
        configured_provider = os.getenv('AI_PROVIDER', '').lower()
        self.provider = configured_provider or ('gemini' if os.getenv('GEMINI_API_KEY') else 'ollama')
        self.api_key = api_key or os.getenv('GEMINI_API_KEY' if self.provider == 'gemini' else 'AI_API_KEY')
        if self.provider == 'gemini':
            self.api_url = api_url or os.getenv(
                'GEMINI_API_URL',
                'https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent'
            )
            self.model = model or os.getenv('GEMINI_MODEL', 'gemini-2.5-flash')
            return
        self.api_url = api_url or os.getenv(
            'AI_API_URL',
            'http://localhost:11434/v1/chat/completions' if self.provider == 'ollama'
            else 'https://api.openai.com/v1/chat/completions'
        )
        self.model = model or os.getenv(
            'AI_MODEL', 'qwen2.5:3b' if self.provider == 'ollama' else 'gpt-4o-mini'
        )

    def generate(self, bot, content_id, question_count):
        self._validate_count(question_count)
        source_text = self._read_all_files(bot, content_id)
        if not source_text.strip():
            raise AIQuizGenerationError(
                'No readable file content was found. A Telegram file may be larger than the 20 MB download limit.'
            )
        if self.provider != 'ollama' and not self.api_key:
            raise AIQuizGenerationError('AI service is not configured.')

        notes = []
        chunks = self._chunk_text(source_text)
        for index, chunk in enumerate(chunks):
            summary = self._complete(self._summary_prompt(chunk))
            if isinstance(summary, dict) and isinstance(summary.get('facts'), list):
                notes.append('\n'.join(str(fact) for fact in summary['facts']))
            else:
                notes.append(json.dumps(summary, ensure_ascii=True))
            
            # فاصل زمني منتظم وثابت (20 ثانية) بين كل جزء والتاني في كل الحالات
            if index < len(chunks) - 1:
                time.sleep(20)

        # استراحة هادئة قبل توليد الأسئلة النهائية
        time.sleep(15)
        questions = self._complete(self._generation_prompt('\n\n'.join(notes), question_count))
        return self._validate_questions(questions, question_count)

    def _read_all_files(self, bot, content_id):
        texts = []
        for file_row in get_content_files(content_id):
            if file_row['file_type'] == 'audio':
                continue
            try:
                file_info = bot.get_file(file_row['telegram_file_id'])
                file_size = getattr(file_info, 'file_size', None)
                if file_size and file_size > self.TELEGRAM_DOWNLOAD_LIMIT:
                    print(
                        f"Skipping AI source file {file_row['file_id']}: "
                        "Telegram download limit is 20 MB."
                    )
                    continue
                data = bot.download_file(file_info.file_path)
                text = self._extract_text(data, file_row['file_name'] or '')
                if text.strip():
                    texts.append(f"FILE: {file_row['file_name'] or 'Unnamed'}\n{text}")
            except Exception as error:
                error_text = str(error).lower()
                if 'file is too big' in error_text or 'too big' in error_text:
                    print(
                        f"Skipping AI source file {file_row['file_id']}: "
                        "Telegram download limit is 20 MB."
                    )
                else:
                    print(f"Could not process AI source file {file_row['file_id']}: {error}")
        return '\n\n'.join(texts)

    def _extract_text(self, data, file_name):
        extension = os.path.splitext(file_name.lower())[1]
        if not extension and data.startswith(b'%PDF'):
            extension = '.pdf'
        if not extension and data.startswith(b'PK'):
            extension = '.docx'
        if extension in {'.txt', '.csv'}:
            return data.decode('utf-8', errors='ignore')
        if extension == '.docx':
            import io
            import docx
            document = docx.Document(io.BytesIO(data))
            return '\n'.join(paragraph.text for paragraph in document.paragraphs)
        if extension == '.pdf':
            import io
            try:
                from pypdf import PdfReader
            except ImportError:
                try:
                    from PyPDF2 import PdfReader
                except ImportError as error:
                    raise AIQuizGenerationError('PDF support is not installed.') from error
            reader = PdfReader(io.BytesIO(data))
            return '\n'.join(page.extract_text() or '' for page in reader.pages)
        if extension in {'.xlsx', '.xls'}:
            try:
                from openpyxl import load_workbook
            except ImportError as error:
                raise AIQuizGenerationError('Spreadsheet support is not installed.') from error
            import io
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            rows = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    rows.append(' | '.join(str(value) for value in row if value is not None))
            return '\n'.join(rows)
        raise AIQuizGenerationError(f'Unsupported source file: {file_name}')

    def _complete(self, prompt):
        if self.provider == 'gemini':
            return self._complete_gemini(prompt)
        payload = {
            'model': self.model,
            'temperature': 0.2,
            'messages': [
                {'role': 'system', 'content': 'You create accurate university MCQs using only supplied source text.'},
                {'role': 'user', 'content': prompt}
            ],
        }
        if self.provider != 'ollama':
            payload['response_format'] = {'type': 'json_object'}
        headers = {'Content-Type': 'application/json'}
        if self.api_key:
            headers['Authorization'] = f'Bearer {self.api_key}'
        request = urllib.request.Request(
            self.api_url,
            data=json.dumps(payload).encode('utf-8'),
            headers=headers,
            method='POST'
        )
        last_error = None
        for attempt in range(3):
            try:
                with urllib.request.urlopen(request, timeout=90) as response:
                    body = json.loads(response.read().decode('utf-8'))
                content = body['choices'][0]['message']['content']
                if isinstance(content, dict):
                    return content
                return json.loads(self._clean_json_response(content))
            except urllib.error.HTTPError as error:
                if error.code == 401:
                    raise AIQuizGenerationError('AI authentication failed. Check the API key and provider URL.') from error
                if error.code == 429:
                    if attempt == 2:
                        raise AIQuizGenerationError('AI service rate limit reached. Please try again later.') from error
                    retry_after = error.headers.get('Retry-After', '5')
                    try:
                        wait_seconds = min(max(int(retry_after), 1), 30)
                    except ValueError:
                        wait_seconds = 5
                    time.sleep(wait_seconds)
                    continue
                last_error = error
            except urllib.error.URLError as error:
                if self.provider == 'ollama':
                    raise AIQuizGenerationError('Local AI is not running. Start Ollama and try again.') from error
                last_error = error
            except (TimeoutError, KeyError, IndexError, json.JSONDecodeError) as error:
                last_error = error
        raise AIQuizGenerationError('The AI returned an invalid response or timed out.') from last_error

    def _complete_gemini(self, prompt):
        payload = {
            'contents': [{'parts': [{'text': prompt}]}],
            'generationConfig': {'temperature': 0.2, 'responseMimeType': 'application/json'}
        }
        endpoint = self.api_url.format(model=self.model)
        separator = '&' if '?' in endpoint else '?'
        request = urllib.request.Request(
            f'{endpoint}{separator}key={urllib.parse.quote(self.api_key)}',
            data=json.dumps(payload).encode('utf-8'),
            headers={'Content-Type': 'application/json'},
            method='POST'
        )
        try:
            with urllib.request.urlopen(request, timeout=90) as response:
                body = json.loads(response.read().decode('utf-8'))
            content = body['candidates'][0]['content']['parts'][0]['text']
            return json.loads(self._clean_json_response(content))
        except urllib.error.HTTPError as error:
            if error.code in (401, 403):
                raise AIQuizGenerationError('Gemini authentication failed. Check the API key.') from error
            if error.code == 429:
                raise AIQuizGenerationError('Gemini free usage limit reached. Please try again later.') from error
            raise AIQuizGenerationError('Gemini could not generate the quiz. Please try again later.') from error
        except (urllib.error.URLError, TimeoutError, KeyError, IndexError, json.JSONDecodeError) as error:
            raise AIQuizGenerationError('Gemini returned an invalid response or timed out.') from error

    def _clean_json_response(self, content):
        content = content.strip()
        if content.startswith('```'):
            content = re.sub(r'^```(?:json)?\s*|\s*```$', '', content, flags=re.IGNORECASE)
        return content.strip()

    def _summary_prompt(self, chunk):
        return (
            'Extract a concise factual study outline from this source text. Preserve only information stated in the text. '
            'Do not add outside knowledge. Return JSON: {"facts":["..."]}.\n\nSOURCE:\n' + chunk
        )

    def _generation_prompt(self, notes, question_count):
        return (
            f'Create exactly {question_count} distinct English university-level multiple-choice questions from the source notes below. '
            'Use strictly only these notes, no internet or outside knowledge. '
            'Requirements: '
            '1. Write everything exclusively in strict English. '
            '2. Focus primarily on deep conceptual understanding and critical thinking rather than surface-level rote memorization, while keeping a balanced mix. '
            '3. Keep questions and options concise and short enough to fit comfortably on a mobile phone screen. '
            '4. Each question must have four short options (1-5 words), one correct answer. '
            'Return JSON only in this exact shape: '
            '{"questions":[{"question":"...","options":{"A":"...","B":"...","C":"...","D":"..."},"correctAnswer":"A"}]}.\n\nSOURCE NOTES:\n'
            + notes
        )

    def _validate_count(self, question_count):
        if question_count not in self.ALLOWED_COUNTS:
            raise AIQuizGenerationError('Question count must be 20, 40, 60, 80, or 100.')

    def _validate_questions(self, response, question_count):
        if not isinstance(response, dict) or not isinstance(response.get('questions'), list):
            raise AIQuizGenerationError('AI response format is invalid.')
        questions = response['questions']
        if len(questions) != question_count:
            raise AIQuizGenerationError('AI did not generate the requested number of questions.')

        seen_questions = set()
        validated = []
        for item in questions:
            question = str(item.get('question', '')).strip()
            options = item.get('options')
            correct = str(item.get('correctAnswer', '')).strip().upper()
            if not question or not isinstance(options, dict) or set(options) != {'A', 'B', 'C', 'D'} or correct not in options:
                raise AIQuizGenerationError('AI returned an invalid question.')
            question_key = re.sub(r'\W+', ' ', question.lower()).strip()
            if question_key in seen_questions:
                raise AIQuizGenerationError('AI returned duplicate questions.')
            seen_questions.add(question_key)
            option_values = [str(options[key]).strip() for key in ('A', 'B', 'C', 'D')]
            if any(not value or len(value.split()) > self.MAX_OPTION_WORDS or len(value) > self.MAX_OPTION_CHARS for value in option_values):
                raise AIQuizGenerationError('AI returned an option that is too long or empty.')
            if len({value.lower() for value in option_values}) != 4 or not re.search(r'[A-Za-z]', question):
                raise AIQuizGenerationError('AI returned invalid or non-English question content.')
            validated.append({'question': question, 'options': dict(zip(('A', 'B', 'C', 'D'), option_values)), 'correctAnswer': correct})
        random.SystemRandom().shuffle(validated)
        for item in validated:
            keys = ['A', 'B', 'C', 'D']
            random.SystemRandom().shuffle(keys)
            old_options = item['options']
            old_correct = item['correctAnswer']
            item['options'] = dict(zip(('A', 'B', 'C', 'D'), [old_options[key] for key in keys]))
            item['correctAnswer'] = ('A', 'B', 'C', 'D')[keys.index(old_correct)]
        return validated

    def _chunk_text(self, text):
        return [text[index:index + self.CHUNK_SIZE] for index in range(0, len(text), self.CHUNK_SIZE)]